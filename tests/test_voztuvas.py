# v1.6 voztuvo testai - 3M mastelio egzamino (2026-08-28) radiniai 1 ir 2:
# (1) Excel ataskaita virsijus lapo ribas skeliama i failus -1/-2/-3
#     (duomenys NEBEDINGSTA; iki v1.5 lapas buvo tyliai apkerpamas ties 1M),
#     jungiklis SDF_EXCEL_ROW_LIMIT leidzia voztuvą pamatyti su 50 eiluciu;
# (2) find_suspects vardo kibiro lubos MAX_NAME_BUCKET - kryzminiai backup
#     medziai nebeuzkabina kvadratinio prefiltro (50 min del 0 poru).
# Self-contained kaip test_engine.py; be Qt.
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import duplicate_engine as de
import exporter
from kalba import t as _t


def _groups(n, dyd=100):
    """n grupiu po 2 netikrus .txt kelius + sizes zodynas (exporter'iui
    disko skaityti nereikia - dydziai paduodami is RAM kaip programoje)."""
    groups, sizes = [], {}
    for i in range(n):
        a = f"C:\\poligonas\\a{i}\\fail{i}.txt"
        b = f"C:\\poligonas\\b{i}\\fail{i}.txt"
        groups.append([a, b])
        sizes[a] = sizes[b] = dyd
    return groups, sizes


def test_be_voztuvo_vienas_failas_be_pastabu(tmp_path, monkeypatch):
    monkeypatch.delenv("SDF_EXCEL_ROW_LIMIT", raising=False)
    groups, sizes = _groups(5)
    out = tmp_path / "report.xlsx"
    info = {}
    p = exporter.export_excel({"groups": groups}, [], out_path=str(out),
                              sizes=sizes, info_out=info)
    assert p == str(out) and out.exists()
    assert info["parts"] == 1 and info["paths"] == [str(out)]
    # normalus atvejis identiskas ankstesnems versijoms: pirmoji eilute -
    # antrastes, jokios "DALIS ..." pastabos
    from openpyxl import load_workbook
    ws = load_workbook(p, read_only=True)[_t("Dublikatai")]
    first = next(ws.iter_rows(values_only=True))
    assert first[0] == _t("Grupe")


def test_voztuvas_skelia_nepraranda_ir_neskaldo_grupiu(tmp_path, monkeypatch):
    monkeypatch.setenv("SDF_EXCEL_ROW_LIMIT", "50")
    groups, sizes = _groups(60)          # 1 seimos antraste + 120 duomenu eil.
    out = tmp_path / "report.xlsx"
    info = {}
    p = exporter.export_excel({"groups": groups}, [], out_path=str(out),
                              sizes=sizes, info_out=info)
    assert info["parts"] >= 3
    assert not out.exists()              # be priedo failo nebera - tik -1/-2/...
    assert Path(info["paths"][0]).name == "report-1.xlsx"
    assert p == info["paths"][0]

    from openpyxl import load_workbook
    keliai_pagal_grupe = {}
    visos_eilutes = 0
    for k, kelias in enumerate(info["paths"], 1):
        wb = load_workbook(kelias, read_only=True)
        assert _t("Dublikatai") in wb.sheetnames
        rows = list(wb[_t("Dublikatai")].iter_rows(values_only=True))
        # pastaba VIRSUJE, pirmoje eiluteje (3M pamoka: apacioje jos nemato)
        laukiama = _t("DALIS {k} IS {n} - eiluciu daugiau nei telpa viename "
                      "Excel faile, ataskaita padalinta i {n} failus").format(
            k=k, n=info["parts"])
        assert rows[0][0] == laukiama
        assert rows[1][0] == _t("Grupe")             # antrastes po pastabos
        assert len(rows) - 2 <= 50                   # lubos gerbiam kiekvienoje dalyje
        for r in rows[2:]:
            if r[2] is None:                         # seimos antrastes eilute
                continue
            visos_eilutes += 1
            keliai_pagal_grupe.setdefault(r[0], set()).add(k)
    # NE VIENA eilute nedingo (voztuvo esme - duomenys nebekerpami)
    assert visos_eilutes == 120
    # ir ne viena grupe nesuskaldyta per du failus
    perskeltos = {g: d for g, d in keliai_pagal_grupe.items() if len(d) > 1}
    assert perskeltos == {}


def test_lubu_jungiklis_prispaudziamas(monkeypatch):
    # per didele ar sugadinta reiksme negali issaudyti ataskaitos
    monkeypatch.setenv("SDF_EXCEL_ROW_LIMIT", "99999999")
    assert exporter._eiluciu_lubos() == exporter.KIETA_RIBA - exporter.ATSARGA
    monkeypatch.setenv("SDF_EXCEL_ROW_LIMIT", "grybas")
    assert exporter._eiluciu_lubos() == exporter.KIETA_RIBA - exporter.ATSARGA
    monkeypatch.setenv("SDF_EXCEL_ROW_LIMIT", "50")
    assert exporter._eiluciu_lubos() == 50


def test_find_suspects_kibiro_lubos_praleidzia_su_pastaba():
    # 2001 vienavardziu failu is "backup medziu" - kibiras virs MAX_NAME_BUCKET
    # praleidziamas PRIES kvadratini prefiltra ir be jokio disko skaitymo
    # (keliai netikri: jei kas nors bandytu hash'uoti, testas kristu letai)
    failai = [(f"C:\\bk{i}\\icon.png", 100 + i % 3)
              for i in range(de.MAX_NAME_BUCKET + 1)]
    stats = {}
    suspects, truncated = de.find_suspects(failai, stats_out=stats)
    assert suspects == [] and truncated is False
    assert stats["skipped_buckets"] == 1
    assert stats["skipped_files"] == de.MAX_NAME_BUCKET + 1


def test_find_suspects_normalus_kelias_stats_nuliai(tmp_path):
    p1 = tmp_path / "report.txt"
    p1.write_bytes(b"a" * 1000)
    sub = tmp_path / "sub"
    sub.mkdir()
    p2 = sub / "report.txt"
    p2.write_bytes(b"b" * 1050)
    stats = {}
    suspects, truncated = de.find_suspects(
        [(str(p1), 1000), (str(p2), 1050)], stats_out=stats)
    assert len(suspects) == 1 and truncated is False
    assert stats == {"skipped_buckets": 0, "skipped_files": 0}
