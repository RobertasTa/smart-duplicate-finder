"""
exporter.py - Excel ataskaita (FAZE 5.2; perrasyta 2026-08-05 greiciui)
Ta pati logika kaip GUI lenteleje: tipo seimu antrastes, grupes atskirtos
dvieju atspalviu kaita. Stulpeliu plociai - automatiskai.

GREICIO sprendimai (dideliems rezultatams, pvz. viso disko skenui):
- write_only Workbook: eilutes srautu i diska, RAM nelaiko viso failo
  (OKF_openpyxl_changelog: write-only rezimas palaiko stilius, plocius,
  bet ws.column_dimensions privalo buti nustatyti PRIES appendinant eilutes);
- sizes zodynas is skeno (RAM) vietoj pakartotiniu os.path.getsize;
- stiliu objektai sukuriami po VIENA karta ir perpanaudojami.
"""
import os
from datetime import datetime
from pathlib import Path

import atranka
from kalba import fam as _famv, t as _t
from table_populator import family_of, FAMILY_ORDER, FAMILY_COLORS


def _rgb(hex_color):
    return "FF" + hex_color.lstrip("#").upper()


MAX_COL_WIDTH = 60  # per placiu stulpeliu lentele istampo; ilgesni tekstai
                    # lauzomi langelio viduje (wrap_text)

# Roberto gyvo testo pastaba 2026-08-24: v1.4 "Kodel" stulpelis nukrisdavo
# UZ EKRANO krasto - matesi tik pirmoji raide. Paaiskinimo, del kurio visa
# funkcija ir daryta, zmogus paprastai net nepamatydavo. Todel du siauresni
# rezimai: kelias 45 vietoj 60, o "Kodel" 26 su lauzymu (eilute paaugsta
# pati, nes Excel su wrap_text auksti pritaiko automatiskai).
KELIO_WIDTH = 45
PRIEZASTIES_WIDTH = 26
# Pastabu stulpelis "Panasiu nuotrauku" lape (2026-08-26). Dvi pastabos
# kartu ("mazesnes raiskos...; kita orientacija...") duoda iki 105 simboliu -
# be lauzymo tai nukristu uz krasto lygiai taip pat, kaip v1.4 "Kodel".
PASTABOS_WIDTH = 34
# Stulpeliu indeksai (0-based)
STULP_KELIAS = 2
STULP_PASTABA = 4      # tik vizualiniame lape; dublikatu lape ten varnele
STULP_KODEL = 5


def _autofit(sheet, rows, headers):
    """Nustato plocius pagal jau paruostus duomenis (vienas praejimas RAM'e).
    Plotis ribojamas - likusi dali sutvarko wrap_text (eilute paaugsta pati)."""
    from openpyxl.utils import get_column_letter
    ribos = {STULP_KELIAS: KELIO_WIDTH, STULP_KODEL: PRIEZASTIES_WIDTH,
             STULP_PASTABA: PASTABOS_WIDTH}
    # Antrastes lauziamos (wrap_text), tad joms uztenka ILGIAUSIO ZODZIO -
    # kitaip "Greiciausiai pirminis" istemptu placia stulpeli varnelei
    widths = [max((len(z) for z in h.split()), default=len(h)) for h in headers]
    for r in rows:
        for i, v in enumerate(r):
            l = len(str(v))
            if l > widths[i]:
                widths[i] = l
    for i, w in enumerate(widths, start=1):
        riba = ribos.get(i - 1, MAX_COL_WIDTH)
        sheet.column_dimensions[get_column_letter(i)].width = min(w + 3, riba)


# --- Excel eiluciu VOZTUVAS (v1.6; Roberto -1/-2/-3 ideja 2026-08-28) ---
# Excel KIETA riba - 1 048 576 eiluciu lape; virsijus Excel "remontuoja" faila
# ismesdamas lapo turini. Iki v1.5 lapa apkirpdavom ties 1M su pastaba
# PABAIGOJE - 3M egzaminas (2026-08-28) parode, kad pastabos 1 000 002-oje
# eiluteje niekas nemato, o duomenys virs ribos DINGSTA. Dabar duomenys
# nebedingsta: ataskaita skeliama i KELIS FAILUS report-1/-2/-3 (ne lapus -
# Excel krauna VISA faila i atminti, atskiri failai atsidaro greiciau ir
# dalinami po viena).
KIETA_RIBA = 1_048_576
ATSARGA = 500  # antrastems ir pastabu eilutems


def _eiluciu_lubos():
    """SDF_EXCEL_ROW_LIMIT - diagnostinis jungiklis testams (antirez principas:
    gudriai logikai atskiras jungiklis, kad vozuva galetum pamatyti ir su 50
    eiluciu, ne tik su milijonu). Gamyboje nenustatytas; didesne uz Excel riba
    reiksme prispaudziama, kad jungikliu nebutu imanoma issaudyti ataskaitos."""
    try:
        v = int(os.environ.get("SDF_EXCEL_ROW_LIMIT", ""))
    except ValueError:
        v = 0
    if v > 0:
        return min(v, KIETA_RIBA - ATSARGA)
    return KIETA_RIBA - ATSARGA


def _dalys(rows, lubos):
    """Skelia eiluciu sarasa i dalis po <= lubos eiluciu, kerpant TIK ties
    grupes riba (pirmo stulpelio reiksmes pasikeitimu), kad viena grupe
    nesiskirstytu per du failus. Seimos antrastes eilute (font be fill)
    klijuojama prie po jos einancios grupes. Jei vienas blokas pats virsija
    lubas (sintetika/testai) - kerpama kietai, kad failas liktu atidaromas."""
    if len(rows) <= lubos:
        return [rows]
    blokai = []
    for r in rows:
        vals, fill, font = r
        if font is not None and fill is None:   # seimos antraste
            blokai.append([r])
            continue
        if blokai:
            pv, pf, pfont = blokai[-1][-1]
            lipdyti = (pfont is not None and pf is None) or pv[0] == vals[0]
        else:
            lipdyti = False
        if lipdyti:
            blokai[-1].append(r)
        else:
            blokai.append([r])
    dalys, dabartine = [], []
    for b in blokai:
        if dabartine and len(dabartine) + len(b) > lubos:
            dalys.append(dabartine)
            dabartine = []
        if len(b) > lubos:
            for i in range(0, len(b), lubos):
                gabalas = b[i:i + lubos]
                if len(gabalas) == lubos:
                    dalys.append(gabalas)
                else:
                    dabartine = gabalas
        else:
            dabartine.extend(b)
    if dabartine:
        dalys.append(dabartine)
    return dalys


def export_excel(scan_results, suspect_results, output_dir=".", out_path=None,
                 sizes=None, visual=None, visual_rotated=None,
                 visual_smaller=None, info_out=None):
    """Export Excel ataskaita su openpyxl (write_only srautas).
    Sheet 'Duplicates': Group | File Name | Full Path | Size (MB),
    rusiuota seima -> grupe, su seimu antrastemis (kaip GUI lenteleje).
    Sheet 'Suspects': same format for suspects.
    Ispejimai: raudona >1GB, geltona >100MB.
    out_path - pilnas tikslo kelias (is 'Kur issaugoti?' dialogo).
    sizes - {kelias: dydis_baitais} is skeno (kad nereiketu getsize is disko)."""

    try:
        from openpyxl import Workbook
        from openpyxl.cell import WriteOnlyCell
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        print("openpyxl nepasiekiamas")
        return None

    # Ilgi keliai lauzomi langelio viduje (Excel auksti pritaiko pats)
    wrap_align = Alignment(wrap_text=True, vertical="top")

    sizes = sizes or {}

    def _size_of(fp):
        s = sizes.get(fp)
        if s is not None:
            return s
        try:
            return os.path.getsize(fp)
        except OSError:
            return 0

    # --- Stiliai: po viena objekta, perpanaudojami ---
    def _fill(hex_color):
        rgb = _rgb(hex_color)
        return PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")

    header_fill = _fill("#4CAF50")
    bold_font = Font(bold=True)
    warn_red = _fill("#FFCDD2")     # >1GB
    warn_yellow = _fill("#FFF9C4")  # >100MB
    fam_fills = {f: (_fill(c[0]), _fill(c[1])) for f, c in FAMILY_COLORS.items()}
    fam_hdr_fonts = {f: Font(bold=True, color=_rgb(c[2]))
                     for f, c in FAMILY_COLORS.items()}
    sus_fills = (_fill("#FFFFCC"), _fill("#FFF3CD"))

    # Roberto klausimas 2026-08-06: lapu pavadinimai ir antrastes - pagal
    # pasirinkta kalba (ataskaita yra dokumentas zmogui), tie patys raktai
    # kaip GUI lenteleje. Failo pavadinimas lieka angliskas (techninis).
    headers = [_t("Grupe"), _t("Failo vardas"),
               _t("Pilnas kelias"), _t("Dydis (MB)")]

    # v1.4: du papildomi stulpeliai TIK dublikatu lape. Tai INFORMACIJA
    # ("kuris cia senelis ir kodel"), NE nurodymas trinti - programa
    # netrina nieko ir siulymo nedaro. Kur pozymiu nera - "neaisku".
    dup_headers = headers + [_t("Greiciausiai pirminis"), _t("Kodel")]
    vis_headers = headers + [_t("Pastaba")]

    # Priezasciu kodai -> tekstai (raktai verciami iprastu keliu)
    _PRIEZ_TEKSTAI = {
        atranka.PRIEZASTIS_VARDAS: _t("kiti grupeje vardu pazymeti kaip kopijos"),
        atranka.PRIEZASTIS_KOPIJU_APLANKAS: _t("kiti guli kopiju aplankuose"),
        atranka.PRIEZASTIS_APLANKAS: _t("kiti guli laikinuose aplankuose"),
        atranka.PRIEZASTIS_GYLIS: _t("kiti guli giliau aplankuose"),
        atranka.PRIEZASTIS_DATA: _t("kiti sukurti veliau"),
        atranka.NEAISKU: _t("neaisku - pozymiu nera"),
    }

    def _pirminis_ir_kodel(grp):
        """Grazina (zyme_siai_eilutei_dict, kodel_tekstas_grupei).
        mtime is disko imamas TIK jei pirmos trys taisykles neisspreze -
        taip dazniausiu atveju disko neliecia is viso."""
        kelias, priez = atranka.greiciausiai_pirminis(grp)
        if priez == atranka.NEAISKU and len(grp) > 1:
            mt = {}
            for fp in grp:
                try:
                    mt[fp] = os.stat(fp).st_mtime
                except OSError:
                    pass
            if mt:
                kelias, priez = atranka.greiciausiai_pirminis(grp, mtimes=mt)
        return kelias, _PRIEZ_TEKSTAI.get(priez, "")

    # --- 1. Paruosiam eiluciu duomenis RAM'e (tik dubliai - nedidele apimtis) ---
    # (vals, fill, font) - fill/font None = paprastas tekstas
    dup_rows = []
    groups = scan_results.get("groups", [])
    fam_groups = {}
    for gid, grp in enumerate(groups, 1):
        if not grp:
            continue
        fam = family_of(Path(grp[0]).suffix)
        fam_groups.setdefault(fam, []).append((gid, grp))

    # PATAISA 2026-08-25: ispejimas "visos sios grupes kopijos guli laikinuose
    # aplankuose" turi prasme tik tada, kai jis ISSKIRIA grupe is kitu.
    # Zmogui, kuris skenuoja butent "Atsisiuntimus" (viena dazniausiu dubliu
    # paieskos priezasciu), jis atsirastu prie KIEKVIENOS grupes ir virstu
    # triuksmu - o kartojamas ispejimas nebera ispejimas. Tokiu atveju
    # nutylim: tai butu ne informacija apie failus, o konstatavimas, kur
    # zmogus pats nusprende ieskoti.
    netuscios = [g for g in groups if g]
    visos_laikinos = bool(netuscios) and all(
        atranka.ZYME_VISI_LAIKINI in atranka.grupes_zymes(g)
        for g in netuscios)

    for fam in FAMILY_ORDER:
        if fam not in fam_groups:
            continue
        exts = sorted({Path(fp).suffix.lower()
                       for _, g in fam_groups[fam] for fp in g})
        # Antraste be fono - tik paryskintas spalvotas tekstas (kaip GUI)
        dup_rows.append(([f"{_famv(fam)} ({', '.join(exts)})", "", "", "", "", ""],
                         None, fam_hdr_fonts[fam]))
        for shade_i, (gid, grp) in enumerate(fam_groups[fam]):
            group_fill = fam_fills[fam][shade_i % 2]
            pirminis, kodel = _pirminis_ir_kodel(grp)
            # Grupes lygio ispejimas (DC "warning all marked" atitikmuo musu
            # kalba): visos sios grupes kopijos guli laikinuose aplankuose
            if (not visos_laikinos
                    and atranka.ZYME_VISI_LAIKINI in atranka.grupes_zymes(grp)):
                zyme = _t("visos kopijos laikinuose aplankuose")
                kodel = f"{kodel}; {zyme}" if kodel else zyme
            pirma_eilute = True
            for fp in grp:
                size_bytes = _size_of(fp)
                if size_bytes > 1_073_741_824:
                    active = warn_red
                elif size_bytes > 104_857_600:
                    active = warn_yellow
                else:
                    active = group_fill
                # "Kodel" rasom VIENA karta grupeje: prie pazymeto failo,
                # o jei nepazymeta nieko - prie pirmos grupes eilutes
                yra_pirminis = bool(pirminis) and fp == pirminis
                if yra_pirminis or (not pirminis and pirma_eilute):
                    kodel_langelis = kodel
                else:
                    kodel_langelis = ""
                dup_rows.append(([_t("Grupe {idx}").format(idx=gid),
                                  os.path.basename(fp),
                                  str(Path(fp).resolve()),
                                  round(size_bytes / 1048576, 2),
                                  "✓" if yra_pirminis else "",
                                  kodel_langelis],
                                 active, None))
                pirma_eilute = False

    # Vizualiai panasios nuotraukos - atskiras lapas (violetiniai atspalviai)
    vis_fills = (_fill("#E9DDF7"), _fill("#CDB4EE"))
    # Pasuktos/veidrodines kopijos zymimos ATSKIRAI (Roberto pastaba
    # 2026-08-26): tai daznai ne kopija, o BROKAS - "pasuko ir pamirso
    # grazinti", ir zmogui verta ta pamatyti pries siunciant krūva klientui.
    pasukti = set(visual_rotated or [])
    mazesni = set(visual_smaller or [])
    zyme_orient = _t("kita orientacija nei kiti grupeje - patikrinkite, ar taip ir turi buti")
    zyme_raiska = _t("mazesnes raiskos nei kiti grupeje")
    vis_rows = []
    for vidx, grp in enumerate(visual or [], 1):
        fill = vis_fills[(vidx - 1) % 2]
        for fp in grp:
            pastabos = []
            if fp in mazesni:
                pastabos.append(zyme_raiska)
            if fp in pasukti:
                pastabos.append(zyme_orient)
            vis_rows.append(([_t("Vaizdas {idx}").format(idx=vidx),
                              os.path.basename(fp),
                              str(Path(fp).resolve()),
                              round(_size_of(fp) / 1048576, 2),
                              "; ".join(pastabos)],
                             fill, None))

    sus_rows = []
    suspects = suspect_results if isinstance(suspect_results, list) else []
    for sid, pair in enumerate(suspects):
        fill = sus_fills[sid % 2]
        for fp, sz_key in ((pair.get("file_a"), "size_a"),
                           (pair.get("file_b"), "size_b")):
            if not fp:
                continue
            # Dydis is variklio (size_a/size_b) - be disko uzklausu
            sz = pair.get(sz_key)
            if sz is None:
                sz = _size_of(fp)
            sus_rows.append(([_t("Itartinas {n}").format(n=sid + 1),
                              os.path.basename(fp),
                              str(Path(fp).resolve()),
                              round(sz / 1048576, 2)],
                             fill, None))

    # --- 2. write_only srautas: plociai PRIES eilutes, tada liejam ---
    # v1.6 VOZTUVAS: virsijus lapo ribas duomenys nebekerpami, o skeliami
    # i kelis failus (zr. _dalys). Normalus atvejis (viena dalis) - failas
    # ir jo turinys IDENTISKI ankstesnems versijoms, jokiu pastabu.

    def _write_sheet(wb, title, rows, lapo_headers=None, top_notice=None):
        # lapo_headers: dublikatu lapas turi dvi papildomas skiltis, kiti - ne
        lapo_headers = lapo_headers or headers
        ws = wb.create_sheet(title)
        # Antrasciu i eiluciu sarasa NEDEDAM - jos jau ateina atskiru
        # argumentu ir skaiciuojamos pagal ilgiausia ZODI (nes lauziamos);
        # idejus dukart laimedavo pilnas ilgis ir stulpelis likdavo platus
        _autofit(ws, [r[0] for r in rows], lapo_headers)
        if top_notice:
            # Pastaba VIRSUJE (3M egzamino pamoka: pastabos apacioje po
            # milijono eiluciu niekas niekada nepamate)
            c = WriteOnlyCell(ws, value=top_notice)
            c.font = bold_font
            ws.append([c])
        hdr = []
        for h in lapo_headers:
            c = WriteOnlyCell(ws, value=h)
            c.fill = header_fill
            c.font = bold_font
            # Ilgas antrastes lauziam - kitaip "Greiciausiai pirminis"
            # isteptu 27 simboliu ploti stulpeliui, kuriame tik varnele
            c.alignment = wrap_align
            hdr.append(c)
        ws.append(hdr)
        for vals, fill, font in rows:
            cells = []
            for ci, v in enumerate(vals):
                c = WriteOnlyCell(ws, value=v)
                if fill is not None:
                    c.fill = fill
                if font is not None:
                    c.font = font
                # Ilgus tekstus lauziam langelio viduje - Excel eilutes
                # auksti tada pritaiko pats (Roberto prasymas 2026-08-24)
                if isinstance(v, str) and (
                        (ci == STULP_KELIAS and len(v) > KELIO_WIDTH)
                        or (ci == STULP_KODEL and len(v) > PRIEZASTIES_WIDTH)
                        or (ci == STULP_PASTABA and len(v) > PASTABOS_WIDTH)):
                    c.alignment = wrap_align
                cells.append(c)
            ws.append(cells)

    lubos = _eiluciu_lubos()
    lapai = [(_t("Dublikatai"), _dalys(dup_rows, lubos), dup_headers)]
    if vis_rows:
        lapai.append((_t("Panasios nuotraukos"), _dalys(vis_rows, lubos),
                      vis_headers))
    lapai.append((_t("Itartini"), _dalys(sus_rows, lubos), headers))
    n_daliu = max(len(d) for _, d, _ in lapai)

    if out_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(output_dir, f"duplicate_report_{ts}.xlsx")

    if n_daliu == 1:
        keliai = [str(out_path)]
    else:
        p = Path(out_path)
        keliai = [str(p.with_name(f"{p.stem}-{k}{p.suffix}"))
                  for k in range(1, n_daliu + 1)]

    for k, kelias in enumerate(keliai, 1):
        wb = Workbook(write_only=True)
        for pavadinimas, sheet_dalys, lapo_headers in lapai:
            if k > 1 and k > len(sheet_dalys):
                continue    # sio lapo eilutes jau tilpo ankstesnese dalyse
            rows_k = sheet_dalys[k - 1] if k <= len(sheet_dalys) else []
            notice = None
            if n_daliu > 1:
                notice = _t("DALIS {k} IS {n} - eiluciu daugiau nei telpa "
                            "viename Excel faile, ataskaita padalinta i {n} "
                            "failus").format(k=k, n=n_daliu)
            _write_sheet(wb, pavadinimas, rows_k, lapo_headers, notice)
        wb.save(kelias)

    if info_out is not None:
        info_out["parts"] = n_daliu
        info_out["paths"] = list(keliai)
    return keliai[0]
